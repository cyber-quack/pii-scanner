#!/usr/bin/env python3
"""
PII Scanner - Detects commonly exposed personally identifiable information in files.

Use Case: Security audits, data leakage prevention, compliance checking.
WARNING: For educational and authorized testing purposes only.
"""

import os
import re
import sys
import warnings
import email
from pathlib import Path
from email import policy
from docx import Document
from openpyxl import load_workbook

# Install dependencies first:
# pip install PyMuPDF python-docx openpyxl
# vvv Suppresses all openpyxl warnings preventing them from appearing in output
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# PII Patterns
PATTERNS = {
    'email': r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}',
    'phone': r'(?:\([2-9]\d{2}\)\s?[2-9]\d{2}[-.]\d{4}|[2-9]\d{2}[-.][2-9]\d{2}[-.]\d{4})',
    'phone_loose': r'(?:\+?1[\s.-]?)?\(?[2-9]\d{2}\)?[-.\s]?[2-9]\d{2}[-.\s]?\d{4}|\b[2-9]\d{9}\b',
    'ssn': r'\b\d{3}-\d{2}-\d{4}\b',
    'credit_card': r'\b(?:4\d{3}|5[1-5]\d{2}|6011|65\d{2})[-\s]?\d{4}[-\s]?\d{4}[-\s]?\d{4}\b',
    'amex': r'\b3[47]\d{2}[-\s]?\d{6}[-\s]?\d{5}\b',
    'credit_card_masked': r'(?:[*xX]{4}[-\s]?){2,3}\d{4}'
}

SCAN_EXTENSIONS = {'.txt', '.py', '.json', '.csv', '.log', '.xml', '.html', '.md', '.ini', '.cfg', '.pdf', '.docx', '.xlsx', '.eml'}

EXCLUDED_DIRS = {'Sys', 'Emulator', 'Games', 'Emu', 'Roms', 'Wads', 'Saves'}

def display_name(pii_type):
    """Map internal pattern names to user-facing category names."""
    mapping = {
        'phone_loose': 'phone',
        'amex': 'credit_card',
        'credit_card_masked': 'credit_card'
    }
    return mapping.get(pii_type, pii_type)
    
def print_banner():
    """Print welcome banner."""
    print("\n" + "=" * 60)
    print("       PII SCANNER v2.0 - Data Privacy Audit Tool")
    print("=" * 60 + "\n")

def get_user_preferences():
    """Interactively collect user preferences for scanning."""
    valid_types = set(PATTERNS.keys()) | {'phone_loose'}
    
    print("Welcome to PII Scanner!")
    print("-" * 60)
    print("\nPII Types to Search:")
    print("  - email       : Email addresses")
    print("  - phone       : Phone numbers")
    print("  - ssn         : Social Security Numbers")
    print("  - credit_card : Credit card numbers")
    
    print("\nInstructions:")
    print("  - Type one or more types, separated by commas")
    print("  - Example: email,phone,ssn")
    print("  - Leave blank to scan for ALL types\n")
    
    # Loop until valid input
    while True:
        pii_input = input("Enter PII types to search: ").strip().lower()
        
        if not pii_input:
            types_to_scan = list(PATTERNS.keys())
            break
        
        types_to_scan = [t.strip().replace(' ', '_') for t in pii_input.split(',')]
        types_to_scan = list(dict.fromkeys(types_to_scan))  # Remove duplicates
        
        invalid_types = [t for t in types_to_scan if t not in valid_types]
        
        if invalid_types:
            print(f"\n[!] Invalid type(s): {invalid_types}")
            print(f"    Valid types: {', '.join(sorted(valid_types - {'phone_loose'}))}")
            print("    Please try again.\n")
            continue
        
        if not types_to_scan:
            print("\n[!] No valid types entered. Please try again.\n")
            continue
        
        break
    
    print("-" * 60)
    print("\nSpecific Values Filter:")
    print("  - Enter comma-separated values to filter results")
    print("  - Example: @outlook.com,555,123-45-6789")
    print("  - Leave blank to show ALL findings\n")
    
    filters = {}
    for pii_type in types_to_scan:
        display_type = pii_type.replace('_', ' ')
        
        filter_val = input(f"Specific {display_type} values (comma-sep, or empty): ").strip()
        if filter_val:
            filters[pii_type] = [v.strip() for v in filter_val.split(',')]
        else:
            filters[pii_type] = None
    
    # When phone is selected, use loose pattern only (it covers both strict and non-strict)
    # Strict matches become DEFINITE, non-strict become POSSIBLE
    if 'phone' in types_to_scan:
        types_to_scan.remove('phone')
        if 'phone_loose' not in types_to_scan:
            types_to_scan.append('phone_loose')
        filters['phone_loose'] = filters.get('phone')
        if 'phone' in filters:
            del filters['phone']
    
    # When credit_card is selected, auto-enable amex (displays merged as CREDIT CARD)
    if 'credit_card' in types_to_scan and 'amex' not in types_to_scan:
        types_to_scan.append('amex')
        filters['amex'] = filters.get('credit_card')
    
    # Same for credit_card_masked - finding redactions means finding card data
    if 'credit_card' in types_to_scan and 'credit_card_masked' not in types_to_scan:
        types_to_scan.append('credit_card_masked')
        filters['credit_card_masked'] = filters.get('credit_card')
    
    return types_to_scan, filters

def get_target_path():
    """Get target directory or file from user."""
    print("\n" + "-" * 60)
    target = input("Enter directory or file path: ").strip()
    
    # Expand ~ and handle Windows paths
    target = os.path.expanduser(target)
    
    while not os.path.exists(target):
        print(f"[!] Error: Path '{target}' does not exist.")
        target = input("Enter a valid directory or file path: ").strip()
        target = os.path.expanduser(target)
    
    return target

def matches_filter(value, filter_list):
    """Check if a found PII value matches any filter value (partial match)."""
    if filter_list is None:
        return True
    for f in filter_list:
        if f.lower() in value.lower():
            return True
        # For phone numbers, compare digits only
        value_digits = re.sub(r'[^0-9]', '', value)
        filter_digits = re.sub(r'[^0-9]', '', f)
        if filter_digits and filter_digits in value_digits:
            return True
    return False

def luhn_check(card_number):
    """Validate a credit card number using the Luhn algorithm."""
    digits = re.sub(r'[-\s]', '', card_number)
    if len(digits) not in (13, 15, 16):
        return False

    total = 0
    reverse = digits[::-1]

    for i, char in enumerate(reverse):
        d = int(char)
        if i % 2 == 1:
            d *= 2
            if d > 9:
                d -= 9
        total += d

    return total % 10 == 0

def validate_credit_card_length(card_number):
    """Ensure credit card has exactly 15 or 16 digits (depending on issuer)."""
    digits = re.sub(r'[-\s]', '', card_number)
    if len(digits) not in (15, 16):
        return False

    if digits.startswith(('34', '37')) and len(digits) != 15:
        return False
    if not digits.startswith(('34', '37')) and len(digits) != 16:
        return False
    return True

def get_phone_confidence(phone_match, strict_pattern_match):
    """Determine confidence level for a phone number match."""
    if strict_pattern_match:
        return 'HIGH'
    # Loose pattern but passes basic NANP validation
    digits = re.sub(r'[^0-9]', '', phone_match)
    if len(digits) >= 10 and digits[:1] in ('2','3','4','5','6','7','8','9'):
        return 'MEDIUM'
    return 'LOW'

def extract_pdf_text(filepath):
    """Extract text content from a PDF file using PyMuPDF."""
    try:
        import pymupdf
        doc = pymupdf.open(filepath)
        text = ""
        for page in doc:
            text += page.get_text()
        doc.close()
        return text
    except Exception as e:
        return None

def extract_docx_text(filepath):
    """Extract text content from a .docx file using python-docx."""
    try:
        doc = Document(str(filepath))
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    except Exception:
        return None

def extract_xlsx_text(filepath):
    """Extract text content from an Excel workbook."""
    try:
        wb = load_workbook(filename=str(filepath), read_only=True)
        text = ""
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        text += str(cell) + "\n"
        wb.close()
        return text
    except Exception:
        return None

def extract_eml_text(filepath):
    """Extract text from .eml files using built-in email module."""
    try:
        with open(filepath, 'rb') as f:
            msg = email.message_from_bytes(f.read(), policy=policy.default)
        text = f"{msg.get('subject', '')}\n{msg.get('from', '')}\n{msg.get('to', '')}\n"
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == 'text/plain':
                    payload = part.get_payload(decode=True)
                    if payload:
                        text += payload.decode(errors='ignore')
        else:
            payload = msg.get_payload(decode=True)
            if payload:
                text += payload.decode(errors='ignore')
        return text
    except Exception:
        return None

def scan_file(filepath, filters):
    """Scan a single file for PII patterns, returning high/low confidence findings."""
    findings_high = {}
    findings_low = {}

    try:
        suffix = filepath.suffix.lower()

        if suffix == '.pdf':
            content = extract_pdf_text(filepath)
            if not content:
                return findings_high, findings_low
        elif suffix == '.docx':
            content = extract_docx_text(filepath)
            if not content:
                return findings_high, findings_low
        elif suffix == '.xlsx':
            content = extract_xlsx_text(filepath)
            if not content:
                return findings_high, findings_low
        elif suffix == '.eml':
            content = extract_eml_text(filepath)
            if not content:
                return findings_high, findings_low
        else:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()

        for pii_type, pattern in PATTERNS.items():
            if pii_type not in filters:
                continue

            filter_list = filters[pii_type]
            matches = re.findall(pattern, content)

            filtered = []
            low_filtered = []

            strict_phone_matches = set()
            if pii_type == 'phone_loose':
                strict_phone_matches = set(re.findall(PATTERNS['phone'], content))

            for match in matches:
                if not matches_filter(match, filter_list):
                    continue

                confidence = 'HIGH'

                if pii_type in ('credit_card', 'amex'):
                    if not luhn_check(match) or not validate_credit_card_length(match):
                        confidence = 'LOW'
                elif pii_type == 'phone_loose':
                    if match in strict_phone_matches:
                        confidence = 'HIGH'
                    else:
                        confidence = 'LOW'
                elif pii_type == 'credit_card_masked':
                    confidence = 'LOW'

                if confidence == 'HIGH':
                    filtered.append((match, 'DEFINITE'))
                else:
                    low_filtered.append((match, 'POSSIBLE'))

            if filtered:
                dtype = display_name(pii_type)
                if dtype not in findings_high:
                    findings_high[dtype] = []
                findings_high[dtype].extend(filtered)

            if low_filtered:
                dtype = display_name(pii_type)
                if dtype not in findings_low:
                    findings_low[dtype] = []
                findings_low[dtype].extend(low_filtered)

    except Exception as e:
        print(f"[!] Error scanning {filepath}: {e}")

    return findings_high, findings_low

def scan_directory(target_path, filters):
    """Recursively scan all files in a directory."""
    path = Path(target_path)
    file_results_high = {}
    file_results_low = {}

    print(f"\n[*] Scanning: {path.resolve()}")
    print("-" * 60)

    for file_path in path.rglob('*'):
        if any(excluded in str(file_path) for excluded in EXCLUDED_DIRS):
            continue

        if file_path.is_file() and file_path.suffix.lower() in SCAN_EXTENSIONS:
            findings_high, findings_low = scan_file(file_path, filters)

            if findings_high:
                rel_path = str(file_path.relative_to(path))
                file_results_high[rel_path] = findings_high
                print(f"[!] Found PII in: {rel_path}")
            
            if findings_low:
                rel_path = str(file_path.relative_to(path))
                file_results_low[rel_path] = findings_low

    return file_results_high, file_results_low

def print_definite_report(file_results):
    """Print high-confidence findings."""
    if not file_results:
        print("\n[+] No DEFINITE PII detected in scanned files.")
        return True

    print("\n" + "=" * 60)
    print("       DEFINITE PII FINDINGS")
    print("=" * 60)

    total_by_type = {}
    file_count = 0
    
    for filename, findings in sorted(file_results.items()):
        file_count += 1
        print(f"\n{'─' * 60}")
        print(f"FILE: {filename}")
        print(f"{'─' * 60}")

        for pii_type, data in sorted(findings.items()):
            unique_values = list(set(data))
            count = len(unique_values)
            
            if pii_type not in total_by_type:
                total_by_type[pii_type] = 0
            total_by_type[pii_type] += count
            
            type_label = pii_type.replace('_', ' ').upper()
            print(f"  [{type_label}] {count} unique value(s)")
            for value, confidence in unique_values:
                print(f"    → {value} [{confidence}]")

    print(f"\n{'=' * 60}")
    print(f"SUMMARY: {file_count} file(s) with DEFINITE PII found")
    for pii_type, count in sorted(total_by_type.items()):
        type_label = pii_type.replace('_', ' ').upper()
        print(f"  {type_label}: {count} unique total")
    print("=" * 60)
    return False

def print_possible_report(file_results):
    """Print low-confidence findings."""
    if not file_results:
        print("\n[+] No POSSIBLE/UNCONFIRMED PII found.")
        return
    
    print("\n" + "=" * 60)
    print("       POSSIBLE/UNCONFIRMED PII FINDINGS")
    print("=" * 60)

    for filename, findings in sorted(file_results.items()):
        print(f"\n{'─' * 60}")
        print(f"FILE: {filename}")
        print(f"{'─' * 60}")

        for pii_type, data in sorted(findings.items()):
            print(f"  [{pii_type.upper()}] {len(data)} instance(s)")
            for value, confidence in data:
                print(f"    → {value} ({confidence})")

def main():
    print_banner()
    
    types_to_scan, filters = get_user_preferences()
    target = get_target_path()

    print("\n[*] Scanning... Please wait.\n")

    if os.path.isfile(target):
        file_results_high, file_results_low = {}, {}
        findings_high, findings_low = scan_file(Path(target), filters)
        if findings_high:
            file_results_high[str(target)] = findings_high
        if findings_low:
            file_results_low[str(target)] = findings_low
    else:
        file_results_high, file_results_low = scan_directory(target, filters)

    # Print definite results first
    has_pii = print_definite_report(file_results_high)
    
    # Optionally show low-confidence results
    if file_results_low:
        print("\n" + "-" * 60)
        show_low = input("Found POSSIBLE/UNCONFIRMED results. Show them? (Y/N): ").strip().upper()
        if show_low == 'Y':
            print_possible_report(file_results_low)

    print("\n[*] Scan complete.")
    print("=" * 60 + "\n")

if __name__ == "__main__":
    main()